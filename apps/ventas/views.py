from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction, models
from decimal import Decimal
from datetime import timedelta, date,timezone
from django.utils import timezone
from datetime import datetime
from apps.ventas.models import Pedido, DetallePedido
from apps.ventas.forms import PedidoForm, TipoDocumentoForm, BoletaForm, FacturaForm
from apps.productos.models import Producto
from apps.documentos.models import DocumentoVenta, DetalleDocumento

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.db.models import Sum, Count

# =========================
# LISTAR PEDIDOS
# =========================
@login_required
def listar_pedidos(request):
    """Muestra todos los pedidos con sus documentos asociados"""
    pedidos = Pedido.objects.select_related('cliente', 'usuario').prefetch_related('detalles').all().order_by('-fecha_creacion')

    # Filtros opcionales por usuario o cliente
    filtro_usuario = request.GET.get('usuario')
    filtro_cliente = request.GET.get('cliente')

    if filtro_usuario:
        pedidos = pedidos.filter(usuario__username__icontains=filtro_usuario)
    if filtro_cliente:
        pedidos = pedidos.filter(cliente__nombre__icontains=filtro_cliente)

    context = {
        'pedidos': pedidos,
    }
    return render(request, 'ventas/listar_pedidos.html', context)


# =========================
@login_required
def crear_pedido_datos(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if request.method == 'POST':
        # Hacemos mutable el POST para modificarlo
        post = request.POST.copy()

        modalidad = post.get('modalidad_pago', 'ahora')

        # Inyectar fecha_emision si modalidad es 'ahora' y no viene en POST
        if modalidad == 'ahora' and not post.get('fecha_emision'):
            post['fecha_emision'] = timezone.localdate().isoformat()            # Limpiar fecha_vencimiento porque no aplica
            post['fecha_vencimiento'] = ''

        pedido_form = PedidoForm(post, instance=pedido)
        tipo_form = TipoDocumentoForm(post)
        boleta_form = BoletaForm(post)
        factura_form = FacturaForm(post)

        if not (pedido_form.is_valid() and tipo_form.is_valid()):
            messages.error(request, 'Revisa los datos del pedido y el tipo de documento.')
            return render(request, 'ventas/crear_pedido_datos.html', {
                'pedido_form': pedido_form,
                'tipo_form': tipo_form,
                'boleta_form': boleta_form,
                'factura_form': factura_form,
                'pedido': pedido,
            })

        # actualiza cliente/observaciones del pedido
        pedido_form.save()

        tipo_doc = tipo_form.cleaned_data['tipo_documento']

        # Crear documento
        doc = DocumentoVenta()
        doc.pedido = pedido
        doc.cliente = pedido.cliente
        doc.tipo_documento = tipo_doc
        doc.vendedor = request.user  # asigna el usuario actual

        # Totales
        neto = sum((dp.subtotal for dp in pedido.detalles.all()), Decimal('0'))
        doc.neto = neto
        doc.iva = (neto * Decimal('0.19')).quantize(Decimal('1.'))
        doc.total = doc.neto + doc.iva

        if tipo_doc == 'Factura':
            if not factura_form.is_valid():
                messages.error(request, 'Revisa los datos de la factura.')
                return render(request, 'ventas/crear_pedido_datos.html', {
                    'pedido_form': pedido_form,
                    'tipo_form': tipo_form,
                    'boleta_form': boleta_form,
                    'factura_form': factura_form,
                    'pedido': pedido,
                })

            # Campos de factura
            doc.razon_social = factura_form.cleaned_data['razon_social']
            doc.rut = factura_form.cleaned_data['rut']
            doc.giro = factura_form.cleaned_data['giro']
            doc.direccion = factura_form.cleaned_data['direccion']
            doc.ciudad = factura_form.cleaned_data['ciudad']
            doc.comuna = factura_form.cleaned_data['comuna']
            doc.medio_de_pago = factura_form.cleaned_data['medio_de_pago']

            dias_plazo = post.get('dias_plazo')

            fecha_emision_user = factura_form.cleaned_data.get('fecha_emision')
            fecha_venc_user = factura_form.cleaned_data.get('fecha_vencimiento')

            if modalidad == 'plazos':
                doc.fecha_emision = fecha_emision_user or date.today()
                
                if fecha_venc_user:
                    doc.fecha_vencimiento = fecha_venc_user
                elif dias_plazo:
                    try:
                        doc.fecha_vencimiento = doc.fecha_emision + timedelta(days=int(dias_plazo))
                    except (ValueError, TypeError):
                        doc.fecha_vencimiento = None
                else:
                    doc.fecha_vencimiento = None
                doc.estado = 'Emitida'
            else:
                # Pagar ahora: fecha de hoy
                doc.fecha_emision = timezone.now()
                doc.fecha_vencimiento = None
                doc.estado = 'Emitida'

        else:
            # Boleta
            if boleta_form.is_valid():
                doc.medio_de_pago = boleta_form.cleaned_data['medio_de_pago']
            doc.fecha_emision = date.today()
            doc.fecha_vencimiento = None
            doc.estado = 'Emitida'

        doc.save()

        # Copiar detalles del pedido al documento
        for d in pedido.detalles.all():
            DetalleDocumento.objects.create(
                documento=doc,
                producto=d.producto,
                cantidad=d.cantidad,
                precio_unitario_venta=d.precio_unitario,
                costo_unitario_venta=getattr(d.producto, 'costo_unitario', Decimal('0')),
            )

        messages.success(request, 'Documento creado correctamente.')
        return redirect('agregar_productos_pedido', pedido_id=pedido.id)

    # GET
    pedido_form = PedidoForm(instance=pedido)
    tipo_form = TipoDocumentoForm()
    boleta_form = BoletaForm()
    factura_form = FacturaForm()

    return render(request, 'ventas/crear_pedido_datos.html', {
        'pedido_form': pedido_form,
        'tipo_form': tipo_form,
        'boleta_form': boleta_form,
        'factura_form': factura_form,
        'pedido': pedido,
    })
# =========================
# PASO 2: AGREGAR PRODUCTOS AL PEDIDO
# =========================
@login_required
def agregar_productos_pedido(request, pedido_id):
    """Vista para agregar productos al pedido y actualizar el documento"""
    
    # Obtener pedido con relaciones precargadas
    pedido = get_object_or_404(
        Pedido.objects.select_related('cliente', 'usuario', 'documentoventa'),
        id=pedido_id
    )
    
    # Verificar que el documento existe
    if not hasattr(pedido, 'documentoventa') or not pedido.documentoventa:
        messages.error(request, "⚠️ Este pedido no tiene un documento asociado. Por favor, contacte al administrador.")
        return redirect('lista_pedidos')
    
    documento = pedido.documentoventa
    
    # Obtener productos disponibles
    productos = Producto.objects.filter(activo=True).order_by('nombre')
    
    # Obtener detalles del pedido (carrito)
    detalles = DetallePedido.objects.filter(pedido=pedido).select_related('producto')
    
    # Preparar carrito para el template
    carrito = []
    for detalle in detalles:
        carrito.append({
            'producto_id': detalle.producto.id,
            'nombre': detalle.producto.nombre,
            'cantidad': detalle.cantidad,
            'precio': detalle.precio_unitario_venta,  # ✅ Campo correcto
            'subtotal': detalle.subtotal,
        })
    
    mensaje_error = None
    
    if request.method == 'POST':
        # ✅ CONFIRMAR PEDIDO
        if 'confirmar_pedido' in request.POST:
            if not detalles.exists():
                messages.error(request, "⚠️ No puede confirmar un pedido sin productos.")
                return redirect('agregar_productos_pedido', pedido_id=pedido.id)
            
            try:
                with transaction.atomic():
                    # Cambiar estado del pedido
                    pedido.estado = 'Enviado'
                    pedido.save()
                    
                    # Actualizar estado del documento
                    documento.estado = 'Emitida'
                    documento.save()
                    
                    # Crear detalles del documento desde los detalles del pedido
                    for detalle in detalles:
                        DetalleDocumento.objects.create(
                            documento=documento,
                            producto=detalle.producto,
                            cantidad=detalle.cantidad,
                            precio_unitario_venta=detalle.precio_unitario_venta,  # ✅ Campo correcto
                            subtotal=detalle.subtotal,
                            costo_unitario_venta=detalle.producto.costo_unitario
                        )
                    
                    messages.success(request, f"✅ Pedido #{pedido.id} confirmado exitosamente.")
                    return redirect('detalle_pedido', pedido_id=pedido.id)
                    
            except Exception as e:
                messages.error(request, f"❌ Error al confirmar el pedido: {str(e)}")
                return redirect('agregar_productos_pedido', pedido_id=pedido.id)
        
        # ✅ AGREGAR PRODUCTO AL PEDIDO
        else:
            producto_id = request.POST.get('producto_id')
            cantidad = request.POST.get('cantidad')
            
            if not producto_id or not cantidad:
                mensaje_error = "⚠️ Debe seleccionar un producto y especificar la cantidad."
            else:
                try:
                    cantidad = int(cantidad)
                    producto = get_object_or_404(Producto, id=producto_id)
                    
                    # Validar stock disponible
                    if producto.stock < cantidad:
                        mensaje_error = f"⚠️ Stock insuficiente. Solo hay {producto.stock} unidades disponibles de {producto.nombre}."
                    else:
                        with transaction.atomic():
                            # Verificar si el producto ya está en el pedido
                            detalle_existente = DetallePedido.objects.filter(
                                pedido=pedido,
                                producto=producto
                            ).first()
                            
                            if detalle_existente:
                                # Actualizar cantidad
                                nueva_cantidad = detalle_existente.cantidad + cantidad
                                
                                if producto.stock < nueva_cantidad:
                                    mensaje_error = f"⚠️ Stock insuficiente. Solo hay {producto.stock} unidades disponibles."
                                else:
                                    detalle_existente.cantidad = nueva_cantidad
                                    detalle_existente.subtotal = detalle_existente.cantidad * detalle_existente.precio_unitario_venta  # ✅ Campo correcto
                                    detalle_existente.save()
                                    
                                    # Descontar stock
                                    producto.stock -= cantidad
                                    producto.save()
                                    
                                    messages.success(request, f"✅ Cantidad actualizada: {producto.nombre} x {nueva_cantidad}")
                            else:
                                # Crear nuevo detalle
                                DetallePedido.objects.create(
                                    pedido=pedido,
                                    producto=producto,
                                    cantidad=cantidad,
                                    precio_unitario_venta=producto.precio_unitario,  # ✅ Campo correcto
                                    subtotal=producto.precio_unitario * cantidad
                                )
                                
                                # Descontar stock
                                producto.stock -= cantidad
                                producto.save()
                                
                                messages.success(request, f"✅ Producto agregado: {producto.nombre} x {cantidad}")
                            
                            # ✅ ACTUALIZAR TOTALES DEL DOCUMENTO
                            actualizar_totales_documento(pedido, documento)
                            
                            return redirect('agregar_productos_pedido', pedido_id=pedido.id)
                            
                except ValueError:
                    mensaje_error = "⚠️ La cantidad debe ser un número válido."
                except Exception as e:
                    mensaje_error = f"❌ Error al agregar el producto: {str(e)}"
    
    context = {
        'pedido': pedido,
        'productos': productos,
        'carrito': carrito,
        'mensaje_error': mensaje_error,
    }
    
    return render(request, 'ventas/agregar_productos_pedido.html', context)


def actualizar_totales_documento(pedido, documento):
    """Función auxiliar para recalcular totales del documento"""
    detalles = DetallePedido.objects.filter(pedido=pedido)
    
    neto = sum(d.subtotal for d in detalles)
    iva = neto * Decimal('0.19')
    total = neto + iva
    
    documento.neto = neto
    documento.iva = iva
    documento.total = total
    documento.save()


@login_required
def eliminar_producto_carrito(request, pedido_id, producto_id):
    """Elimina un producto del pedido y restaura el stock"""
    pedido = get_object_or_404(Pedido, id=pedido_id)
    producto = get_object_or_404(Producto, id=producto_id)
    
    try:
        with transaction.atomic():
            detalle = DetallePedido.objects.filter(pedido=pedido, producto=producto).first()
            
            if detalle:
                # Restaurar stock
                producto.stock += detalle.cantidad
                producto.save()
                
                # Eliminar detalle
                detalle.delete()
                
                # Actualizar totales del documento
                if hasattr(pedido, 'documentoventa') and pedido.documentoventa:
                    actualizar_totales_documento(pedido, pedido.documentoventa)
                
                messages.success(request, f"✅ Producto eliminado: {producto.nombre}")
            else:
                messages.warning(request, "⚠️ El producto no está en el pedido.")
                
    except Exception as e:
        messages.error(request, f"❌ Error al eliminar el producto: {str(e)}")
    
    return redirect('agregar_productos_pedido', pedido_id=pedido_id)
# =========================
# DETALLE DEL PEDIDO
# =========================
@login_required
def detalle_pedido(request, pedido_id):
    """Muestra el detalle de un pedido con sus productos y documento asociado"""
    pedido = get_object_or_404(
        Pedido.objects.select_related('cliente', 'usuario')
        .prefetch_related('detalles__producto'),
        id=pedido_id
    )
    
    # Obtener detalles del pedido
    detalles = pedido.detalles.all()
    
    # Verificar si el pedido tiene documento
    try:
        documento = DocumentoVenta.objects.get(pedido=pedido)
    except DocumentoVenta.DoesNotExist:
        documento = None

    context = {
        'pedido': pedido,
        'detalles': detalles,
        'documento': documento,
    }
    return render(request, 'ventas/detalle_pedido.html', context)


# =========================
# CONFIRMAR PEDIDO
# =========================
@login_required
def confirmar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.estado != 'Pendiente':
        messages.warning(request, 'Este pedido ya fue confirmado o procesado.')
        return redirect('detalle_pedido', pedido_id=pedido.id)

    detalles = DetallePedido.objects.filter(pedido=pedido)

    if not detalles.exists():
        messages.error(request, 'No se puede confirmar un pedido sin productos.')
        return redirect('detalle_pedido', pedido_id=pedido.id)

    datos_doc = request.session.get('pedido_temp_data', {})

    try:
        with transaction.atomic():
            # 🔍 Verificar stock suficiente antes de descontar
            productos_sin_stock = []
            for detalle in detalles:
                producto = detalle.producto
                if detalle.cantidad > producto.stock:
                    productos_sin_stock.append(
                        f"{producto.nombre} (Solicitado: {detalle.cantidad}, Disponible: {producto.stock})"
                    )

            if productos_sin_stock:
                mensaje_error = "⚠️ No hay suficiente stock para los siguientes productos:\n" + "\n".join(productos_sin_stock)
                messages.error(request, mensaje_error)
                return redirect('detalle_pedido', pedido_id=pedido.id)

            # 🔻 Descontar stock de cada producto
            for detalle in detalles:
                producto = detalle.producto
                producto.stock -= detalle.cantidad
                producto.save()

            # 💰 Calcular totales
            subtotal = sum(detalle.cantidad * detalle.producto.precio_unitario for detalle in detalles)
            iva = subtotal * Decimal('0.19')
            total = subtotal + iva

            # 🧾 Crear documento de venta
            tipo_documento = datos_doc.get('tipo_documento', 'Boleta')
            
            documento = DocumentoVenta.objects.create(
                pedido=pedido,
                tipo_documento=tipo_documento,
                cliente=pedido.cliente,
                vendedor=request.user,
                medio_de_pago=datos_doc.get('medio_de_pago', ''),
                neto=subtotal,
                iva=iva,
                total=total,
                fecha_vencimiento=datos_doc.get('fecha_vencimiento') if tipo_documento == 'Factura' else None,
                razon_social=datos_doc.get('razon_social', '') if tipo_documento == 'Factura' else '',
                rut=datos_doc.get('rut', '') if tipo_documento == 'Factura' else '',
                giro=datos_doc.get('giro', '') if tipo_documento == 'Factura' else '',
                direccion=datos_doc.get('direccion', '') if tipo_documento == 'Factura' else '',
                ciudad=datos_doc.get('ciudad', '') if tipo_documento == 'Factura' else '',
                comuna=datos_doc.get('comuna', '') if tipo_documento == 'Factura' else '',
            )

            # 📦 Crear detalles del documento
            for detalle in detalles:
                DetalleDocumento.objects.create(
                    documento=documento,
                    producto=detalle.producto,
                    cantidad=detalle.cantidad,
                    precio_unitario_venta=detalle.producto.precio_unitario,
                    costo_unitario_venta=detalle.producto.costo_unitario
                )

            # ✅ Actualizar estado del pedido
            pedido.estado = 'Procesando'
            pedido.save()

            # 💾 Limpiar datos temporales de sesión
            if 'pedido_temp_data' in request.session:
                del request.session['pedido_temp_data']

            messages.success(request, f'El Pedido #{pedido.id} ha sido confirmado. Documento {tipo_documento} #{documento.folio} generado correctamente.')

    except Exception as e:
        messages.error(request, 'Ocurrió un error inesperado al confirmar el pedido.')
        print(e)

    return redirect('detalle_pedido', pedido_id=pedido.id)


# =========================
# MARCAR PEDIDO COMO ENVIADO
# =========================
@login_required
def marcar_pedido_enviado(request, pedido_id):
    """Cambia el estado del pedido a 'Enviado' si está procesando."""
    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.estado != 'Procesando':
        messages.warning(request, 'Solo se pueden marcar como enviados los pedidos en proceso.')
        return redirect('detalle_pedido', pedido_id=pedido.id)

    pedido.estado = 'Enviado'
    pedido.save()

    messages.success(request, f'El pedido #{pedido.id} ha sido marcado como Enviado.')
    return redirect('detalle_pedido', pedido_id=pedido.id)


@login_required
def estadisticas_ventas(request):
    if request.user.rol != 'Administrador':
        messages.error(request, "⚠️ No tienes permisos para acceder a esta sección.")
        return redirect('dashboard')

    # Obtener pedidos enviados con documento precargado
    pedidos = (Pedido.objects.filter(estado='Enviado')
                .select_related('cliente', 'usuario', 'documentoventa')
                .order_by('-fecha_creacion'))

    # Filtros por fecha
    fecha_desde = request.GET.get('fecha_desde')
    fecha_hasta = request.GET.get('fecha_hasta')

    if fecha_desde:
        pedidos = pedidos.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        pedidos = pedidos.filter(fecha_creacion__date__lte=fecha_hasta)

    # Calcular totales
    total_ventas = pedidos.count()
    monto_total = sum(
        p.documentoventa.total 
        for p in pedidos 
        if hasattr(p, 'documentoventa') and p.documentoventa and p.documentoventa.total
    )

    context = {
        'pedidos': pedidos,
        'total_ventas': total_ventas,
        'monto_total': monto_total,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }

    return render(request, 'ventas/estadisticas_ventas.html', context)


@login_required
def exportar_ventas_excel(request):
    if request.user.rol != 'Administrador':
        messages.error(request, "⚠️ No tienes permisos para exportar.")
        return redirect('dashboard')

    pedidos = (Pedido.objects.filter(estado='Enviado')
                .select_related('cliente', 'usuario', 'documentoventa')
                .order_by('-fecha_creacion'))

    fecha_desde_str = request.GET.get('fecha_desde')
    fecha_hasta_str = request.GET.get('fecha_hasta')

    if fecha_desde_str:
        try:
            fecha_desde = datetime.strptime(fecha_desde_str, "%Y-%m-%d").date()
            pedidos = pedidos.filter(fecha_creacion__date__gte=fecha_desde)
        except ValueError:
            messages.error(request, "⚠️ Fecha desde inválida. Usa formato YYYY-MM-DD.")

    if fecha_hasta_str:
        try:
            fecha_hasta = datetime.strptime(fecha_hasta_str, "%Y-%m-%d").date()
            pedidos = pedidos.filter(fecha_creacion__date__lte=fecha_hasta)
        except ValueError:
            messages.error(request, "⚠️ Fecha hasta inválida. Usa formato YYYY-MM-DD.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ['Pedido #', 'Cliente', 'RUT', 'Vendedor', 'Fecha', 'Estado', 'Total']
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for pedido in pedidos:
        total = pedido.documentoventa.total if hasattr(pedido, 'documentoventa') and pedido.documentoventa else 0

        ws.append([
            pedido.id,
            pedido.cliente.razon_social,
            pedido.cliente.rut,
            pedido.usuario.get_full_name() or pedido.usuario.username,
            pedido.fecha_creacion.strftime('%d/%m/%Y %H:%M'),
            pedido.estado,
            f"${total:,.0f}"
        ])

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"ventas_{date.today().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response

def crear_pedido_inicial(request):
    if request.method == 'POST':
        form = PedidoForm(request.POST)
        if form.is_valid():
            pedido = form.save(commit=False)
            pedido.estado = 'Borrador'
            pedido.usuario = request.user  
            pedido.save()
            messages.success(request, f'Pedido #{pedido.id} creado. Ahora define el tipo de documento.')
            return redirect('crear_pedido_datos', pedido_id=pedido.id)
    else:
        form = PedidoForm()
    return render(request, 'ventas/crear_pedido_inicial.html', {'form': form})